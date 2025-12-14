import os
import threading
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook, load_workbook

from .models import MessageBatch, Message
from .forms import ExcelUploadForm
from .whatsapp_bot import WhatsAppBot


# Global reference to the bot thread
_bot_thread = None
_bot_instance = None


def upload_view(request):
    """Main page with Excel upload form."""
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            return process_upload(request, form)
    else:
        form = ExcelUploadForm()
    
    # Get recent batches for display
    recent_batches = MessageBatch.objects.order_by('-created_at')[:10]
    
    return render(request, 'messenger/upload.html', {
        'form': form,
        'recent_batches': recent_batches,
    })


def process_upload(request, form):
    """Process the uploaded Excel file and create messages."""
    excel_file = request.FILES['excel_file']
    attachment_files = request.FILES.getlist('attachments')
    
    # Get recent batches for error page display
    recent_batches = MessageBatch.objects.order_by('-created_at')[:10]
    
    # Parse Excel file first to validate attachments
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        return render(request, 'messenger/upload.html', {
            'form': form,
            'error': f'Error reading Excel file: {str(e)}',
            'recent_batches': recent_batches,
        })
    
    # Collect all attachment filenames from Excel (case-insensitive)
    required_attachments = set()
    excel_rows = []  # Store parsed rows for later use
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:  # Skip empty rows
            continue
            
        phone = str(row[0]).strip() if row[0] else None
        message_text = str(row[1]).strip() if len(row) > 1 and row[1] else None
        attachment_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
        
        if not phone or not message_text:
            continue
        
        excel_rows.append({
            'phone': phone,
            'message': message_text,
            'attachment': attachment_name,
            'row_num': row_num,
        })
        
        # Track required attachments
        if attachment_name:
            required_attachments.add(attachment_name.lower())
    
    # Get uploaded file names (case-insensitive)
    uploaded_filenames = set()
    for f in attachment_files:
        # Handle files from folders - extract just the filename, not the path
        filename = os.path.basename(f.name)
        uploaded_filenames.add(filename.lower())
    
    # Check for missing attachments
    missing_attachments = required_attachments - uploaded_filenames
    
    if missing_attachments:
        # Find which rows reference missing files for detailed error
        missing_details = []
        for row in excel_rows:
            if row['attachment'] and row['attachment'].lower() in missing_attachments:
                missing_details.append({
                    'row': row['row_num'],
                    'phone': row['phone'],
                    'filename': row['attachment'],
                })
        
        error_msg = f"Missing {len(missing_attachments)} attachment(s) referenced in Excel"
        
        return render(request, 'messenger/upload.html', {
            'form': form,
            'error': error_msg,
            'missing_attachments': list(missing_attachments),
            'missing_details': missing_details,
            'uploaded_files': [os.path.basename(f.name) for f in attachment_files],
            'recent_batches': recent_batches,
        })
    
    # Validation passed - now save files and create batch
    try:
        # Create upload directory for this batch
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'uploads', timezone.now().strftime('%Y%m%d_%H%M%S'))
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save attachment files
        attachment_paths = {}
        for f in attachment_files:
            filename = os.path.basename(f.name)
            file_path = os.path.join(upload_dir, filename)
            with open(file_path, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk)
            attachment_paths[filename.lower()] = file_path
        
        # Create the batch
        batch = MessageBatch.objects.create(
            excel_filename=excel_file.name,
            status='pending'
        )
        
        messages_created = 0
        
        # Create messages from parsed rows
        for row in excel_rows:
            attachment_path = None
            if row['attachment']:
                attachment_path = attachment_paths.get(row['attachment'].lower())
            
            Message.objects.create(
                batch=batch,
                phone_number=row['phone'],
                message_text=row['message'],
                attachment_filename=row['attachment'],
                attachment_path=attachment_path,
                status='pending'
            )
            messages_created += 1
        
        # Update batch with total count
        batch.total_messages = messages_created
        batch.save()
        
        return redirect('status', batch_id=batch.id)
        
    except Exception as e:
        return render(request, 'messenger/upload.html', {
            'form': form,
            'error': f'Error processing upload: {str(e)}',
            'recent_batches': recent_batches,
        })


def status_view(request, batch_id):
    """Display the status of a message batch."""
    batch = get_object_or_404(MessageBatch, id=batch_id)
    messages = batch.messages.all().order_by('id')
    
    return render(request, 'messenger/status.html', {
        'batch': batch,
        'messages': messages,
    })


@require_http_methods(["GET"])
def api_status(request, batch_id):
    """API endpoint for getting batch status (for AJAX updates)."""
    batch = get_object_or_404(MessageBatch, id=batch_id)
    messages = batch.messages.all().order_by('id')
    
    return JsonResponse({
        'batch': {
            'id': batch.id,
            'status': batch.status,
            'total_messages': batch.total_messages,
            'sent_count': batch.sent_count,
            'failed_count': batch.failed_count,
            'progress_percent': batch.progress_percent,
        },
        'messages': [
            {
                'id': m.id,
                'phone': m.phone_number,
                'status': m.status,
                'error': m.error_message,
                'has_attachment': bool(m.attachment_filename),
            }
            for m in messages
        ]
    })


@require_http_methods(["POST"])
def start_sending(request, batch_id):
    """Start the WhatsApp sending process for a batch."""
    global _bot_thread, _bot_instance
    
    batch = get_object_or_404(MessageBatch, id=batch_id)
    
    if batch.status == 'running':
        return JsonResponse({'error': 'Batch is already running'}, status=400)
    
    # Update batch status
    batch.status = 'running'
    batch.save()
    
    # Start the bot in a background thread
    def run_bot():
        global _bot_instance
        try:
            # Check if we can reuse the existing bot instance
            if _bot_instance and _bot_instance.is_active():
                print("Reusing existing WhatsApp Bot session")
                # Navigate to home to ensure we are in a known state (and to verify login works)
                try:
                    _bot_instance.driver.get("https://web.whatsapp.com")
                except Exception:
                    # If navigation fails, the driver might be dead despite is_active check
                    _bot_instance = WhatsAppBot()
                    _bot_instance.start()
            else:
                # Create and start a new bot instance
                _bot_instance = WhatsAppBot()
                _bot_instance.start()
            
            # Wait for user to scan QR code (or confirm already logged in)
            if not _bot_instance.wait_for_login(timeout=120):
                batch.status = 'failed'
                batch.save()
                return
            
            # Process each pending message
            messages = batch.messages.filter(status='pending')
            
            for message in messages:
                message.status = 'sending'
                message.save()
                
                success, error = _bot_instance.send_message(
                    message.phone_number,
                    message.message_text,
                    message.attachment_path
                )
                
                if success:
                    message.status = 'sent'
                    message.sent_at = timezone.now()
                    batch.sent_count += 1
                else:
                    message.status = 'failed'
                    message.error_message = error
                    batch.failed_count += 1
                
                message.save()
                batch.save()
                
                # Random delay between messages
                _bot_instance.random_delay()
            
            # Mark batch as completed
            batch.status = 'completed'
            batch.save()
            
        except Exception:
            batch.status = 'failed'
            batch.save()
        finally:
            # Only close the browser if the batch failed or was stopped
            # If completed successfully, keep the browser open as requested
            if batch.status != 'completed' and _bot_instance:
                _bot_instance.close()
                _bot_instance = None
            elif batch.status == 'completed':
                # Keep _bot_instance active so we don't lose the reference
                # (though starting a new batch will overwrite it and potentially fail due to profile lock)
                pass
    
    _bot_thread = threading.Thread(target=run_bot, daemon=True)
    _bot_thread.start()
    
    return JsonResponse({'success': True, 'message': 'Sending started. Please scan the QR code.'})


@require_http_methods(["POST"])
def stop_sending(request, batch_id):
    """Stop the WhatsApp sending process."""
    global _bot_instance
    
    batch = get_object_or_404(MessageBatch, id=batch_id)
    
    if _bot_instance:
        _bot_instance.close()
        _bot_instance = None
    
    # Mark remaining pending messages as failed
    batch.messages.filter(status__in=['pending', 'sending']).update(
        status='failed',
        error_message='Stopped by user'
    )
    
    batch.status = 'failed'
    batch.failed_count = batch.messages.filter(status='failed').count()
    batch.save()
    
    return JsonResponse({'success': True})


def download_template(request):
    """Download a sample Excel template."""
    from django.http import HttpResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"
    
    # Headers
    ws['A1'] = 'Phone Number'
    ws['B1'] = 'Message'
    ws['C1'] = 'Attachment (optional)'
    
    # Sample data
    ws['A2'] = '5512345678'
    ws['B2'] = 'Hello! This is a test message.'
    ws['C2'] = 'document.pdf'
    
    ws['A3'] = '5587654321'
    ws['B3'] = 'Hi there!\nThis message has multiple lines.'
    ws['C3'] = ''
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=whatsapp_template.xlsx'
    
    wb.save(response)
    return response
